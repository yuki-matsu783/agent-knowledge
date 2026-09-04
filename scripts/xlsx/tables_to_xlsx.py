"""markdown の表 または CSV から xlsx を生成する。

使い方:
    uv run python scripts/xlsx/tables_to_xlsx.py <input.md|input.csv ...> -o <output.xlsx>

- .md  : 本文中の GFM 表をすべて拾い、直前の見出しをシート名にする (同名は連番)
- .csv : ファイル名をシート名にする (UTF-8、BOM 可)
- 書式は Anthropic の xlsx skill の要件に合わせる: Arial、太字ヘッダー、先頭行固定、オートフィルタ、列幅
- 数式は書かない (値のみ)。数式が要る場合は openpyxl で別途書き、LibreOffice で再計算する
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="0F3D5E")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT)
INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")

Table = tuple[str, list[list[str]]]  # (シート名候補, 行)


def split_md_row(line: str) -> list[str]:
    body = line.strip()
    if body.startswith("|"):
        body = body[1:]
    if body.endswith("|"):
        body = body[:-1]
    # エスケープされた \| は区切りとして扱わない
    cells = re.split(r"(?<!\\)\|", body)
    return [c.replace("\\|", "|").strip() for c in cells]


def is_separator(line: str) -> bool:
    return bool(re.fullmatch(r"\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)*\|?\s*", line))


def strip_md_inline(text: str) -> str:
    text = re.sub(r"`([^`]*)`", r"\1", text)
    text = re.sub(r"\[([^\]]*)\]\([^)]*\)", r"\1", text)
    text = re.sub(r"\*\*([^*]*)\*\*", r"\1", text)
    return text


def tables_from_markdown(path: Path) -> list[Table]:
    lines = path.read_text(encoding="utf-8").splitlines()
    tables: list[Table] = []
    heading = path.stem
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("#"):
            heading = line.lstrip("#").strip() or heading
        if "|" in line and i + 1 < len(lines) and is_separator(lines[i + 1]):
            rows = [split_md_row(line)]
            i += 2
            while i < len(lines) and "|" in lines[i] and lines[i].strip():
                rows.append(split_md_row(lines[i]))
                i += 1
            tables.append((heading, [[strip_md_inline(c) for c in r] for r in rows]))
            continue
        i += 1
    return tables


def tables_from_csv(path: Path) -> list[Table]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        rows = [list(r) for r in csv.reader(fh)]
    return [(path.stem, rows)]


def coerce(value: str) -> str | int | float:
    if re.fullmatch(r"-?\d+", value):
        return int(value)
    if re.fullmatch(r"-?\d+\.\d+", value):
        return float(value)
    return value


def unique_sheet_name(name: str, used: set[str]) -> str:
    base = INVALID_SHEET_CHARS.sub(" ", name).strip()[:28] or "Sheet"
    candidate = base
    n = 2
    while candidate in used:
        candidate = f"{base[:25]} ({n})"
        n += 1
    used.add(candidate)
    return candidate


def write_sheet(wb: Workbook, name: str, rows: list[list[str]]) -> None:
    ws = wb.create_sheet(title=name)
    width = max((len(r) for r in rows), default=0)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx in range(1, width + 1):
            raw = row[c_idx - 1] if c_idx - 1 < len(row) else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=coerce(raw) if r_idx > 1 else raw)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(vertical="center")
            else:
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=len(raw) > 60)
    for c_idx in range(1, width + 1):
        longest = max((len(str(r[c_idx - 1])) for r in rows if c_idx - 1 < len(r)), default=8)
        # 全角を幅 2 として概算し、上限を設ける
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(10, longest * 1.1 + 2), 60)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(width)}{len(rows)}"


def main(argv: list[str]) -> int:
    # Windows のコンソールが cp932 でも日本語ログを壊さない
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("inputs", nargs="+", type=Path, help="markdown または CSV")
    parser.add_argument("-o", "--output", required=True, type=Path, help="出力 xlsx")
    args = parser.parse_args(argv)

    tables: list[Table] = []
    for path in args.inputs:
        if not path.exists():
            print(f"error: {path} が無い", file=sys.stderr)
            return 1
        if path.suffix.lower() == ".csv":
            tables.extend(tables_from_csv(path))
        elif path.suffix.lower() in {".md", ".markdown"}:
            tables.extend(tables_from_markdown(path))
        else:
            print(f"error: {path} は .md か .csv にする", file=sys.stderr)
            return 1
    if not tables:
        print("error: 表が 1 つも見つからない", file=sys.stderr)
        return 1

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for name, rows in tables:
        write_sheet(wb, unique_sheet_name(name, used), rows)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"wrote: {args.output} sheets={len(tables)}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
